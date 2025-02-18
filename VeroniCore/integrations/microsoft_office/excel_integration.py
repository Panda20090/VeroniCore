# excel_integration.py
# This script handles integration with Microsoft Excel for VeroniCore.
# It provides functionality to read, write, and manipulate Excel files.

import openpyxl

class ExcelIntegration:
    def __init__(self, file_path):
        self.file_path = file_path
        self.workbook = None

    def load_workbook(self):
        try:
            self.workbook = openpyxl.load_workbook(self.file_path)
            print(f"Workbook '{self.file_path}' loaded successfully.")
        except Exception as e:
            print(f"Error loading workbook: {e}")

    def create_workbook(self):
        try:
            self.workbook = openpyxl.Workbook()
            print("New workbook created successfully.")
        except Exception as e:
            print(f"Error creating workbook: {e}")

    def save_workbook(self, save_path=None):
        try:
            save_path = save_path or self.file_path
            self.workbook.save(save_path)
            print(f"Workbook saved successfully at '{save_path}'.")
        except Exception as e:
            print(f"Error saving workbook: {e}")

    def add_sheet(self, sheet_name):
        try:
            self.workbook.create_sheet(title=sheet_name)
            print(f"Sheet '{sheet_name}' added successfully.")
        except Exception as e:
            print(f"Error adding sheet: {e}")

    def remove_sheet(self, sheet_name):
        try:
            sheet = self.workbook[sheet_name]
            self.workbook.remove(sheet)
            print(f"Sheet '{sheet_name}' removed successfully.")
        except Exception as e:
            print(f"Error removing sheet: {e}")

    def write_data(self, sheet_name, cell, data):
        try:
            sheet = self.workbook[sheet_name]
            sheet[cell] = data
            print(f"Data written to '{sheet_name}' at cell '{cell}': {data}")
        except Exception as e:
            print(f"Error writing data: {e}")

    def read_data(self, sheet_name, cell):
        try:
            sheet = self.workbook[sheet_name]
            data = sheet[cell].value
            print(f"Data read from '{sheet_name}' at cell '{cell}': {data}")
            return data
        except Exception as e:
            print(f"Error reading data: {e}")
            return None

if __name__ == "__main__":
    # Example usage
    excel_integration = ExcelIntegration("example.xlsx")
    excel_integration.create_workbook()
    excel_integration.add_sheet("Sheet1")
    excel_integration.write_data("Sheet1", "A1", "Hello, Excel!")
    excel_integration.save_workbook("example.xlsx")
    